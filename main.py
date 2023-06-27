from selenium import webdriver
from selenium.webdriver.common.by import By
from io import BytesIO

from pytesseract import pytesseract
from PIL import Image
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

import base64


import time


proxy_server = '171.234.58.26:25373'
path_to_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.tesseract_cmd = path_to_tesseract

class UserInfo:
    def __init__(self):
        self.user = 'test256147'
        self.pwd = 'Matkhau@2023'
        self.money_pwd = '123456'
        self.name = 'NGUYEN VAN MANH'


def init_driver(proxy_server):

    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument('--proxy-server=' + proxy_server)
    driver = webdriver.Chrome(options=chrome_options)

    return driver

def init_webs(file_path):
    webs = []
    f = open(file_path)
    for line in f:
        webs.append(line)
    
    return webs
def init_user_info(file_path):
    users_info = []
    dataframe = openpyxl.load_workbook(file_path)
    dataframe1 = dataframe.active
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            print(col[row].value)
def fill_register_form(driver, user_info):
    inputs = driver.find_elements(By.TAG_NAME, 'input')
    for input in inputs:
        if input.get_attribute('ng-model') == '$ctrl.user.account.value':
            input.clear()
            input.send_keys(user_info.user)
        if input.get_attribute('ng-model') == '$ctrl.user.password.value':
            input.clear()
            input.send_keys(user_info.pwd)
        if input.get_attribute('ng-model') == '$ctrl.user.confirmPassword.value':
            input.clear()
            input.send_keys(user_info.pwd)
        if input.get_attribute('ng-model') == '$ctrl.user.moneyPassword.value':
            input.clear()
            input.send_keys(user_info.money_pwd)
        if input.get_attribute('ng-model') == '$ctrl.user.name.value':
            input.clear()
            input.send_keys(user_info.name)
        if input.get_attribute('ng-model') == '$ctrl.code':
            input.clear()
            input.click()
            time.sleep(1)
            imgs = driver.find_elements(By.TAG_NAME, 'img')
            for img in imgs:
                if img.get_attribute('ng-class') == '$ctrl.styles.captcha':
                    base64_str = img.get_attribute('src')
                    base64_img = base64_str.split(',')
                    base64_img = base64_img[1]
                    imgdata = base64.b64decode(base64_img)
                    img = Image.open(BytesIO(imgdata))
                    text = pytesseract.image_to_string(img)
                    text = text[0:4]
                    input.send_keys(text)
                    submit = driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
                    submit.click()
                    time.sleep(1000000000)

def open_register_form(driver, url_register):
    driver.get(url_register)
    time.sleep(2)

    spans = WebDriverWait(driver, timeout=3).until(lambda d: d.find_elements(By.TAG_NAME, "span")) 
    for span in spans:
        if span.get_attribute('ng-click') == '$ctrl.ok()':
            span.click()
            time.sleep(1)
            buttons = driver.find_elements(By.TAG_NAME, 'button')
            for button in buttons:
                if button.get_attribute('ng-class') == '$ctrl.styles.reg':
                    button.click()
                    return True


if __name__ == "__main__":
    driver = init_driver(proxy_server)
    time.sleep(2)
    webs = init_webs('webs.txt')
    for url_web in webs:
        is_openned_register_form = open_register_form(driver, url_web)
        time.sleep(2)
        if is_openned_register_form:
            fill_register_form(driver, user_info)
